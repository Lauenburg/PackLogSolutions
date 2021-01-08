import json
from collections import OrderedDict
from itertools import islice
from openpyxl import load_workbook
from datetime import date, datetime

def read_from_xl(path_to_xl, sheet_name, ord_list_column_names):
    """ Passes the sheets of an excel file to a list format.

        Args:
            path_to_xl: Path to the excel file.
            sheet_name: The name of the sheet with in the excel file.
            ord_list_column_names: A ordred list with the names for the columns.

        Return:
            A list of ordered dictionaries containing the information form the excel sheet.
    """

    # Open the workbook and select a worksheet
    wb = load_workbook(path_to_xl)
    sheet = wb[sheet_name]

    # List to hold dictionaries
    dic_list = []

    # Iterate through each row in worksheet and fetch values into dict
    for row in islice(sheet.values, 1, sheet.max_row):
        sub_dic = OrderedDict()
        for i, filed in enumerate(ord_list_column_names):    
            sub_dic[filed] = row[i]
        dic_list.append(sub_dic)
    
    return dic_list

def write_dic_to_json(dic_list, path_save, xl_name, sheet_name):
    """ Writes the a list of dictionaries to a .json file.

        Args:
            dic_list: List containing the dictionaries.
            path_save: Location to save the .json file.
            xl_name: Name of the original excel file.
            sheet_name: Name of the excel file sheet from which the data originates. 

        Return:
            The json string.
    """
    # Serialize the list of dicts to JSON
    j = json.dumps(dic_list, default=json_serial)

    # Write to file
    with open(path_save + sheet_name + "_" + xl_name + '.json', 'w') as f:
        f.write(j)

    return j

def json_serial(obj):
    """JSON serializer for objects not serializable by default json code"""

    if isinstance(obj, (datetime, date)):
        return obj.isoformat()
    raise TypeError ("Type %s not serializable" % type(obj))